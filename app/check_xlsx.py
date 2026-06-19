import openpyxl
from pathlib import Path

xlsx = Path(r"E:\경공매 프로그램\기준시가,공동주택,시가표준액 조회\시가표준액_부산\오피스텔_표준가격기준액_2026.xlsx")
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
print("시트:", wb.sheetnames)

ws = wb.active

# 전체 헤더 확인
headers = [ws.cell(1, c).value for c in range(1, 15)]
print("헤더:", headers)

# 처음 5개 행 확인
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
    print(f"Row{i}:", list(row[:15]))

# 삼정(7047번) 근처 확인
for row in ws.iter_rows(min_row=7040, max_row=7060, values_only=True):
    print("  ", list(row[:6]))
